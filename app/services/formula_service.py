"""Service de génération de formules de parfum.

Charge les données du coffret (XLSX) une seule fois en mémoire,
puis génère 2 formules personnalisées à partir des réponses utilisateur.
"""

import json
from collections import defaultdict
from pathlib import Path

import openpyxl

from app.data.choice_profile_mapping import (
    CHOICE_PROFILE_MAPPING,
    INGREDIENT_EN_TO_FR,
    PROFILE_DESCRIPTIONS,
    PROFILE_DESCRIPTIONS_EN,
    PROFILE_GENDERS,
)
from app.data.questions import EN_TO_FR_CHOICES
from app.services import redis_service

# ── Chemins ───────────────────────────────────────────────────────────
_DATA_DIR = Path(__file__).resolve().parent.parent / "data"
_XLSX_PATH = _DATA_DIR / "Coffret-description.xlsx"

# ── Cache mémoire (chargé une seule fois) ─────────────────────────────
_coffret: dict | None = None


# ── Normalisation des noms de profils ─────────────────────────────────
# Le XLSX utilise des variantes (Stratégist, Disrupteur, Trail blazer…)
_PROFILE_NORMALIZE = {
    "stratégist": "Strategist",
    "strategist": "Strategist",
    "disrupteur": "Disruptor",
    "disruptor": "Disruptor",
    "trail blazer": "Trailblazer",
    "trailblazer": "Trailblazer",
    "visionary": "Visionary",
    "visionnary": "Visionary",
    "innovator": "Innovator",
    "creator": "Creator",
    "influencer": "Influencer",
    "icon": "Icon",
    "cosy": "Cosy",
}


def _normalize_profile(raw: str | None) -> str | None:
    if not raw:
        return None
    return _PROFILE_NORMALIZE.get(raw.strip().lower(), raw.strip())


# ── Chargement du XLSX ────────────────────────────────────────────────

def _load_coffret() -> dict:
    """Parse le XLSX et retourne les ingrédients + allergènes."""
    wb = openpyxl.load_workbook(_XLSX_PATH, read_only=True, data_only=True)

    # --- Sheet 1 : ingrédients ---
    ws = wb[wb.sheetnames[0]]
    ingredients = []

    # Lignes des ingrédients : T=7-16, C=21-30, F=35-44
    note_ranges = [
        ("top", 7, 16),
        ("heart", 21, 30),
        ("base", 35, 44),
    ]

    for note_type, start, end in note_ranges:
        for row in ws.iter_rows(min_row=start, max_row=end, min_col=1, max_col=8):
            position = row[0].value
            if not position:
                continue
            ingredients.append({
                "position": position,
                "name": row[1].value,
                "family": row[2].value,
                "description": row[3].value,
                "note_type": note_type,
                "profile_1": _normalize_profile(row[6].value),
                "profile_2": _normalize_profile(row[7].value),
            })

    # --- Sheet 2 : allergènes ---
    ws2 = wb["ALLERGENS"]

    # 3 blocs d'allergènes : top (row 4-31), heart (row 33-60), base (row 62-89)
    allergen_blocks = [
        (4, 6, 31),   # header_row, data_start, data_end
        (33, 35, 60),
        (62, 64, 89),
    ]

    # allergen_map: ingredient_name → set of allergen names
    allergen_map: dict[str, set[str]] = defaultdict(set)

    for header_row, data_start, data_end in allergen_blocks:
        # Read header: col B onwards = ingredient names
        header_cells = list(ws2.iter_rows(
            min_row=header_row, max_row=header_row,
            min_col=2, max_col=11,
        ))[0]
        col_to_ingredient = {}
        for cell in header_cells:
            if cell.value:
                col_to_ingredient[cell.column] = cell.value.strip()

        # Read allergen rows
        for row in ws2.iter_rows(
            min_row=data_start, max_row=data_end,
            min_col=1, max_col=11,
        ):
            allergen_name = row[0].value
            if not allergen_name:
                continue
            allergen_name = allergen_name.strip()
            for cell in row[1:]:
                if cell.value and str(cell.value).strip().lower() == "x":
                    ingredient_name = col_to_ingredient.get(cell.column)
                    if ingredient_name:
                        allergen_map[ingredient_name].add(allergen_name)

    wb.close()

    return {
        "ingredients": ingredients,
        "allergen_map": dict(allergen_map),
    }


def _get_coffret() -> dict:
    global _coffret
    if _coffret is None:
        _coffret = _load_coffret()
    return _coffret


# ── Scoring des profils ──────────────────────────────────────────────

def _score_profiles(answers: dict, gender: str | None) -> list[tuple[str, float]]:
    """Calcule le score de chaque profil à partir des réponses.

    - top_2 choice → +2 pts aux profils associés
    - bottom_2 choice → -1 pt aux profils associés
    - Bonus genre : +1 si le profil match le genre, +0.5 pour unisex
    """
    scores: dict[str, float] = defaultdict(float)

    for qid_str, answer_data in answers.items():
        qid = int(qid_str)
        mapping = CHOICE_PROFILE_MAPPING.get(qid, {})

        if isinstance(answer_data, str):
            answer_data = json.loads(answer_data)

        en_to_fr = EN_TO_FR_CHOICES.get(qid, {})

        for choice in answer_data.get("top_2", []):
            fr_choice = en_to_fr.get(choice, choice)
            profiles = mapping.get(fr_choice, [])
            for p in profiles:
                scores[p] += 2.0

        for choice in answer_data.get("bottom_2", []):
            fr_choice = en_to_fr.get(choice, choice)
            profiles = mapping.get(fr_choice, [])
            for p in profiles:
                scores[p] -= 1.0

    # Bonus genre
    if gender:
        gender_lower = gender.lower()
        # Map "masculin"/"féminin" (FR) to "masculine"/"feminine"
        if gender_lower in ("masculin", "masculine"):
            target = "masculine"
        elif gender_lower in ("féminin", "feminine", "feminin"):
            target = "feminine"
        else:
            target = None

        if target:
            for profile, profile_gender in PROFILE_GENDERS.items():
                if profile_gender == target:
                    scores[profile] += 1.0
                elif profile_gender == "unisex":
                    scores[profile] += 0.5

    # Trier par score décroissant
    return sorted(scores.items(), key=lambda x: x[1], reverse=True)


# ── Sélection des ingrédients ─────────────────────────────────────────

def _get_ingredients_for_profile(
    profile_name: str,
    user_allergens: list[str] | None = None,
) -> dict[str, list[dict]]:
    """Retourne les ingrédients du coffret associés à un profil.

    Filtre les allergènes si l'utilisateur en a déclaré.
    Profile 1 match → priorité haute, Profile 2 → priorité basse.
    """
    coffret = _get_coffret()
    allergen_map = coffret["allergen_map"]

    # Normaliser les allergènes utilisateur pour comparaison insensible à la casse
    blocked_ingredients: set[str] = set()
    if user_allergens:
        user_allergens_lower = {a.strip().lower() for a in user_allergens}
        for ingredient_name, allergens in allergen_map.items():
            allergens_lower = {a.lower() for a in allergens}
            if allergens_lower & user_allergens_lower:
                blocked_ingredients.add(ingredient_name)

    result: dict[str, list[dict]] = {
        "top_notes": [],
        "heart_notes": [],
        "base_notes": [],
    }

    note_type_key = {"top": "top_notes", "heart": "heart_notes", "base": "base_notes"}

    for ingredient in coffret["ingredients"]:
        if ingredient["name"] in blocked_ingredients:
            continue

        is_p1 = ingredient["profile_1"] == profile_name
        is_p2 = ingredient["profile_2"] == profile_name

        if is_p1 or is_p2:
            key = note_type_key[ingredient["note_type"]]
            result[key].append({
                "position": ingredient["position"],
                "name": ingredient["name"],
                "family": ingredient["family"],
                "description": ingredient["description"],
                "priority": "primary" if is_p1 else "secondary",
            })

    # Trier : primary d'abord
    for key in result:
        result[key].sort(key=lambda x: 0 if x["priority"] == "primary" else 1)

    return result


# ── Génération des formules ───────────────────────────────────────────

def generate_formulas(session_id: str) -> dict:
    """Génère 2 formules personnalisées pour une session.

    1. Récupère réponses + profil depuis Redis
    2. Score les profils
    3. Sélectionne les ingrédients pour les 2 meilleurs profils
    4. Retourne les formules
    """
    # Récupérer les données
    session_data = redis_service.get_session_answers(session_id)
    if not session_data or not session_data.get("answers"):
        return {"error": "Aucune réponse trouvée", "formulas": []}

    # Déterminer la langue de la session
    session_meta = redis_service.get_session_meta(session_id)
    language = session_meta.get("language", "fr") if session_meta else "fr"
    descriptions = PROFILE_DESCRIPTIONS_EN if language == "en" else PROFILE_DESCRIPTIONS
    translate_name = (lambda name: INGREDIENT_EN_TO_FR.get(name, name)) if language == "fr" else (lambda name: name)

    profile = redis_service.get_user_profile(session_id)
    gender = profile.get("gender") if profile else None
    has_allergies = profile.get("has_allergies", "non") if profile else "non"
    user_allergens_raw = profile.get("allergies", "") if profile else ""

    # Parser les allergènes
    user_allergens = None
    if has_allergies == "oui" and user_allergens_raw:
        user_allergens = [a.strip() for a in user_allergens_raw.replace(",", ";").split(";") if a.strip()]

    # Scorer les profils
    ranked = _score_profiles(session_data["answers"], gender)

    if len(ranked) < 2:
        return {"error": "Pas assez de données pour générer des formules", "formulas": []}

    # Prendre les 2 meilleurs, en s'assurant qu'ils sont différents
    top_profiles = []
    for profile_name, score in ranked:
        if len(top_profiles) >= 2:
            break
        top_profiles.append((profile_name, score))

    # Générer les formules
    formulas = []
    for profile_name, score in top_profiles:
        ingredients = _get_ingredients_for_profile(profile_name, user_allergens)
        # Traduire les noms d'ingrédients dans les détails
        translated_details = {}
        for note_key in ("top_notes", "heart_notes", "base_notes"):
            translated_details[note_key] = [
                {**n, "name": translate_name(n["name"])} for n in ingredients[note_key]
            ]

        formulas.append({
            "profile": profile_name,
            "description": descriptions.get(profile_name, ""),
            "score": score,
            "top_notes": [n["name"] for n in translated_details["top_notes"]],
            "heart_notes": [n["name"] for n in translated_details["heart_notes"]],
            "base_notes": [n["name"] for n in translated_details["base_notes"]],
            "details": translated_details,
        })

    return {"formulas": formulas}
